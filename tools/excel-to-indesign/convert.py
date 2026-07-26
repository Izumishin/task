#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
保証人住所リスト Excel -> InDesign データ結合用ファイル 変換ツール

やること
  1. 全角英数字 (０-９ Ａ-Ｚ ａ-ｚ) を半角へ
  2. ハイフン類 (－ ― ‐ – — − および数字に挟まれた長音符 ー) を半角ハイフン - へ統一
     ※ カタカナ語の長音符 (コーポ, タワー 等) は変換しない
  3. 前後の空白除去・連続空白の圧縮
  4. 住所1行目の先頭に都道府県名が重複している場合は除去
  5. Excel が日付に化けない形式で書き出し
     - UTF-16LE(BOM) タブ区切り .txt … InDesign 推奨 / Excel でも安全に開ける
     - UTF-8(BOM) 全項目クオート .csv … 汎用
     - 文字列セルとして保存した .xlsx … Excel 確認用
  6. 目視確認が必要な行を「要確認リスト」に書き出し

使い方:
    python3 convert.py <入力.xlsx> [出力ディレクトリ]
"""

import sys, os, re, csv, unicodedata, collections
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

SHEET_FALLBACK = 0

# ---------------------------------------------------------------- 正規化ルール

# 長音符ではない、明確な「ハイフン類」。どの位置にあっても半角ハイフンにする。
HYPHEN_LIKE = "－―‐–—−﹣－‒⁃"
# 長音符。数字・英字に挟まれているときだけハイフン扱いにする。
PROLONGED = "ーｰ"

_HYPHEN_TABLE = {ord(c): "-" for c in HYPHEN_LIKE}
# 数字/英字 に挟まれた長音符
_PROLONGED_AS_HYPHEN = re.compile(r"(?<=[0-9A-Za-z])[" + PROLONGED + r"](?=[0-9A-Za-z])")
_WS_RUN = re.compile(r"[\s　]+")


def to_hankaku_alnum(s: str) -> str:
    """全角英数字のみを半角化する。記号・カナ・漢字は触らない。"""
    out = []
    for ch in s:
        o = ord(ch)
        if 0xFF10 <= o <= 0xFF19 or 0xFF21 <= o <= 0xFF3A or 0xFF41 <= o <= 0xFF5A:
            out.append(chr(o - 0xFEE0))
        else:
            out.append(ch)
    return "".join(out)


def normalize_spaces(s: str) -> str:
    """前後の空白を除去し、連続空白を1文字に圧縮する。
    全角が含まれる連続空白は全角スペース、それ以外は半角スペースに寄せる。"""
    s = s.strip().strip("　").strip()

    def _rep(m):
        return "　" if "　" in m.group(0) else " "

    return _WS_RUN.sub(_rep, s)


def unify_hyphens(s: str) -> str:
    s = s.translate(_HYPHEN_TABLE)
    s = _PROLONGED_AS_HYPHEN.sub("-", s)
    return s


def normalize_cell(value, unify_hyphen: bool) -> str:
    s = "" if value is None else str(value)
    s = to_hankaku_alnum(s)
    if unify_hyphen:
        s = unify_hyphens(s)
    s = normalize_spaces(s)
    return s


# ---------------------------------------------------------------- 検証ルール

PREFECTURES = set("""北海道 青森県 岩手県 宮城県 秋田県 山形県 福島県 茨城県 栃木県 群馬県
埼玉県 千葉県 東京都 神奈川県 新潟県 富山県 石川県 福井県 山梨県 長野県 岐阜県 静岡県 愛知県
三重県 滋賀県 京都府 大阪府 兵庫県 奈良県 和歌山県 鳥取県 島根県 岡山県 広島県 山口県 徳島県
香川県 愛媛県 高知県 福岡県 佐賀県 長崎県 熊本県 大分県 宮崎県 鹿児島県 沖縄県""".split())

# Excel が自動変換してしまう値のパターン
RE_DATE_YMD = re.compile(r"\d{1,4}[-/]\d{1,2}[-/]\d{1,2}")
RE_DATE_MD = re.compile(r"\d{1,4}[-/]\d{1,2}")
RE_NUM = re.compile(r"\d+")
RE_DECIMAL = re.compile(r"\d+\.\d+")


def excel_autoconvert_kind(s: str):
    """この文字列を Excel が General 書式のセルに入れたとき、何に化けるか。"""
    t = s.strip()
    if not t:
        return None
    if RE_DATE_YMD.fullmatch(t):
        return "日付 (年-月-日と誤認)"
    if RE_DATE_MD.fullmatch(t):
        return "日付 (月-日と誤認)"
    if RE_DECIMAL.fullmatch(t):
        return "小数値"
    if RE_NUM.fullmatch(t):
        return "数値 (先頭0が消える)"
    return None


# 機種依存文字 (ローマ数字・丸数字・単位記号など)
RE_KISHU = re.compile(r"[Ⅰ-ⅿ①-⓿㈠-㉃㊀-㋿㌀-㏿℃℉Å㎜-㎡]")
# 半角化後も残る全角記号
RE_ZEN_SYMBOL = re.compile(r"[！-／：-＠［-｀｛-･〜～]")
RE_BUILDING = re.compile(r"[ァ-ヴ]{3,}|マンション|ハイツ|コーポ|号室|ビル|号棟|階建|レジデンス|ハウス|パレス|タワー")
# 「字クンネベツ」のような小字名はカタカナでも建物名ではない
RE_AZA = re.compile(r"^(大字|小字|字)")
# 法人・団体名の保証人 (敬称が「様」ではなく「御中」になる可能性)
RE_ORG = re.compile(r"機構|株式会社|有限会社|財団|社団|法人|大学|学園|センター|協会|組合|教務課|連携課")


# ---------------------------------------------------------------- 出力

def write_tab_utf16(path, header, rows):
    """UTF-16LE(BOM) タブ区切り。InDesign のデータ結合が最も安定して読む形式で、
    Excel で開いてもテキストインポートウィザードが立つため日付化しない。"""
    with open(path, "w", encoding="utf-16-le", newline="") as f:
        f.write("﻿")
        f.write("\t".join(header) + "\r\n")
        for r in rows:
            f.write("\t".join(r) + "\r\n")


def write_csv_utf8bom(path, header, rows):
    """UTF-8(BOM) / 全項目クオート。InDesign・一般ツール向け。"""
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, quoting=csv.QUOTE_ALL, lineterminator="\r\n")
        w.writerow(header)
        w.writerows(rows)


def write_xlsx_text(path, header, rows, sheet_title="変換後"):
    """全セルを文字列書式 (@) で保存。Excel で開いても日付に化けない確認用。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(header)
    for c in range(1, len(header) + 1):
        ws.cell(1, c).font = Font(bold=True)
    for r in rows:
        ws.append(r)
    for c in range(1, len(header) + 1):
        col = get_column_letter(c)
        for cell in ws[col]:
            cell.number_format = "@"
            cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[col].width = max(10, min(34, max(len(str(x[c - 1])) for x in [header] + rows) * 2 + 2))
    ws.freeze_panes = "A2"
    wb.save(path)


def write_issue_xlsx(path, issues):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "要確認"
    head = ["元Excelの行", "分類", "列", "現在の値", "内容 / 推奨対応"]
    ws.append(head)
    for c in range(1, len(head) + 1):
        ws.cell(1, c).font = Font(bold=True)
    for it in issues:
        ws.append(it)
    for c, w in zip("ABCDE", (12, 26, 18, 34, 62)):
        ws.column_dimensions[c].width = w
        for cell in ws[c]:
            cell.number_format = "@"
            cell.alignment = Alignment(vertical="center", wrap_text=(c == "E"))
    ws.freeze_panes = "A2"
    wb.save(path)


# ---------------------------------------------------------------- メイン

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    src = sys.argv[1]
    outdir = sys.argv[2] if len(sys.argv) > 2 else os.path.dirname(os.path.abspath(src))
    os.makedirs(outdir, exist_ok=True)

    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb.worksheets[SHEET_FALLBACK]
    header = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    ncol = len(header)

    # ハイフン統一を適用する列 (郵便番号・住所)
    hyphen_cols = {i for i, h in enumerate(header) if ("郵便番号" in h or "住所" in h)}
    pref_col = next((i for i, h in enumerate(header) if "都道府県" in h), None)
    addr1_col = next((i for i, h in enumerate(header) if "住所1" in h), None)

    raw_rows, out_rows = [], []
    for r in range(2, ws.max_row + 1):
        raw = [str(ws.cell(r, c).value or "") for c in range(1, ncol + 1)]
        if all(v.strip() == "" for v in raw):
            continue
        raw_rows.append((r, raw))
        out_rows.append([normalize_cell(raw[c], c in hyphen_cols) for c in range(ncol)])

    issues = []
    changelog = []
    stats = collections.Counter()

    def add_issue(excel_row, kind, col_idx, value, note):
        issues.append([str(excel_row), kind, header[col_idx] if col_idx is not None else "-", value, note])
        stats[kind] += 1

    # --- 都道府県の重複除去 + 各種チェック
    zip_pref = collections.defaultdict(collections.Counter)
    for (excel_row, raw), out in zip(raw_rows, out_rows):
        if pref_col is not None:
            zp = next((i for i, h in enumerate(header) if "郵便番号" in h), None)
            if zp is not None:
                zip_pref[out[zp][:3]][out[pref_col]] += 1

    for (excel_row, raw), out in zip(raw_rows, out_rows):
        # 都道府県の重複
        if pref_col is not None and addr1_col is not None:
            pref = out[pref_col]
            if pref and out[addr1_col].startswith(pref):
                before = out[addr1_col]
                out[addr1_col] = before[len(pref):]
                add_issue(excel_row, "都道府県の重複を自動除去", addr1_col, before,
                          f"「{pref}」＋「{before}」で県名が二重になるため先頭を削除 → 「{out[addr1_col]}」")

        # 変更ログ
        for c in range(ncol):
            if raw[c] != out[c]:
                changelog.append([str(excel_row), header[c], raw[c], out[c]])

        # Excel 自動変換リスク (情報として記録)
        for c in range(ncol):
            k = excel_autoconvert_kind(out[c])
            if k and "郵便番号" not in header[c]:
                stats[f"[参考] Excelで{k}になり得るセル"] += 1

        # 都道府県名の妥当性
        if pref_col is not None and out[pref_col] not in PREFECTURES:
            add_issue(excel_row, "都道府県名が不正", pref_col, out[pref_col], "47都道府県のいずれにも一致しません")

        # 郵便番号の書式
        zp = next((i for i, h in enumerate(header) if "郵便番号" in h), None)
        if zp is not None and out[zp] and not re.fullmatch(r"\d{3}-\d{4}", out[zp]):
            add_issue(excel_row, "郵便番号の書式が不正", zp, out[zp], "NNN-NNNN 形式ではありません")

        # 空欄
        for c in range(ncol):
            if out[c] == "" and "住所3" not in header[c]:
                add_issue(excel_row, "必須項目が空欄", c, "(空欄)", "データ結合時に空欄で出力されます")

        # 機種依存文字
        for c in range(ncol):
            m = RE_KISHU.search(out[c])
            if m:
                add_issue(excel_row, "機種依存文字", c, out[c],
                          f"「{m.group(0)}」は環境依存文字です。InDesign のフォントによっては字形が出ません")

        # 半角化後も残る全角記号
        for c in range(ncol):
            found = "".join(sorted(set(RE_ZEN_SYMBOL.findall(out[c]))))
            if found:
                add_issue(excel_row, "全角記号が残存", c, out[c],
                          f"全角記号「{found}」。和文として残すか半角にするか要判断のため自動変換していません")

        # 住所2行目に建物名が混在
        a2 = next((i for i, h in enumerate(header) if "住所2" in h), None)
        a3 = next((i for i, h in enumerate(header) if "住所3" in h), None)
        if a2 is not None and RE_BUILDING.search(out[a2]) and not RE_AZA.match(out[a2]):
            add_issue(excel_row, "住所2行目に建物名が混在", a2, out[a2],
                      "番地欄に建物名が入っています。住所3行目へ分けると組版が揃います")

        # 番地欄の書式崩れ
        if a2 is not None:
            v = out[a2]
            if re.search(r"[0-9][\s　]+[0-9]", v) or re.search(r"[\s　]-|-[\s　]", v):
                add_issue(excel_row, "番地の区切りが不正", a2, v, "数字の間に空白が入っています。ハイフンに直してください")
            if "." in v:
                add_issue(excel_row, "番地に小数点", a2, v, "「.」が使われています。ハイフンまたは全角読点の誤りの可能性")

        # 氏名の区切り
        for c in range(ncol):
            if "氏名" in header[c] and out[c] and not re.search(r"[\s　]", out[c]):
                add_issue(excel_row, "氏名に姓名の区切りなし", c, out[c],
                          "姓と名の間に全角スペースがありません。組版で姓名が続けて出ます")
        # 保証人が法人・団体
        gc = next((i for i, h in enumerate(header) if h == "保証人氏名"), None)
        if gc is not None and RE_ORG.search(out[gc]):
            add_issue(excel_row, "保証人が法人・団体名", gc, out[gc],
                      "個人名ではありません。敬称を「様」ではなく「御中」等にするか個別対応が必要です")

    # 郵便番号3桁 vs 都道府県の突き合わせ (データ内での多数決)
    zp = next((i for i, h in enumerate(header) if "郵便番号" in h), None)
    if zp is not None and pref_col is not None:
        for (excel_row, raw), out in zip(raw_rows, out_rows):
            pre = out[zp][:3]
            cnt = zip_pref.get(pre)
            if not cnt or len(cnt) < 2:
                continue
            main_pref, main_n = cnt.most_common(1)[0]
            if out[pref_col] != main_pref and cnt[out[pref_col]] <= 1 and main_n >= 2:
                add_issue(excel_row, "郵便番号と都道府県の不一致", pref_col, f"{out[zp]} / {out[pref_col]}",
                          f"郵便番号 {pre}-xxxx は他 {main_n} 件すべて「{main_pref}」です。住所「{out[addr1_col]}」と照合してください")

    # 文字数が最大値に張り付いている = 元システムで切れている疑い
    for c in range(ncol):
        lens = [len(o[c]) for o in out_rows if o[c]]
        if not lens:
            continue
        mx = max(lens)
        at_max = [(er, o[c]) for (er, _), o in zip(raw_rows, out_rows) if len(o[c]) == mx]
        if mx < 10 or len(at_max) > 10:
            continue
        # 同じ文字列は1件にまとめる
        by_val = collections.defaultdict(list)
        for er, v in at_max:
            by_val[v].append(er)
        for v, ers in by_val.items():
            # 語尾がカタカナ/ひらがなの途中で終わっている = 切れている可能性が高い
            if not re.search(r"[ァ-ヴぁ-ん]$", v):
                continue
            rows_txt = "、".join(map(str, ers[:6])) + ("ほか" if len(ers) > 6 else "")
            add_issue(ers[0], "文字数上限で切れている疑い", c, v,
                      f"この列の最大文字数 {mx} 文字ちょうどで語尾が不自然です（該当行: {rows_txt}）。元データを確認してください")

    # 重複行
    seen = collections.defaultdict(list)
    for (excel_row, raw), out in zip(raw_rows, out_rows):
        seen[tuple(out)].append(excel_row)
    for k, v in seen.items():
        if len(v) > 1:
            add_issue(v[0], "完全重複行", 0, k[0], f"行 {', '.join(map(str, v))} が完全に同一です")

    # 学生氏名の重複 (同姓同名 or 二重登録)
    sc = next((i for i, h in enumerate(header) if h == "学生氏名"), None)
    if sc is not None:
        by_student = collections.defaultdict(list)
        for (er, _), o in zip(raw_rows, out_rows):
            by_student[o[sc]].append(er)
        for name, ers in by_student.items():
            if len(ers) > 1:
                add_issue(ers[0], "学生氏名の重複", sc, name,
                          f"行 {', '.join(map(str, ers))} に同じ氏名があります。同姓同名か二重登録かご確認ください")

    issues.sort(key=lambda x: (int(x[0]), x[1]))

    base = "保証人住所"
    p_txt = os.path.join(outdir, f"{base}_InDesign_UTF16タブ区切り.txt")
    p_csv = os.path.join(outdir, f"{base}_InDesign_UTF8.csv")
    p_xlsx = os.path.join(outdir, f"{base}_変換後_Excel確認用.xlsx")
    p_iss = os.path.join(outdir, f"{base}_要確認リスト.xlsx")
    p_log = os.path.join(outdir, f"{base}_変換ログ.csv")

    write_tab_utf16(p_txt, header, out_rows)
    write_csv_utf8bom(p_csv, header, out_rows)
    write_xlsx_text(p_xlsx, header, out_rows)
    write_issue_xlsx(p_iss, issues)
    with open(p_log, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, quoting=csv.QUOTE_ALL, lineterminator="\r\n")
        w.writerow(["元Excelの行", "列", "変換前", "変換後"])
        w.writerows(changelog)

    print(f"入力       : {src}")
    print(f"データ行数 : {len(out_rows)}")
    print(f"変換セル数 : {len(changelog)}")
    print(f"要確認件数 : {len(issues)}")
    print("\n--- 内訳 ---")
    for k, n in sorted(stats.items(), key=lambda x: -x[1]):
        print(f"  {k}: {n}")
    print("\n--- 出力 ---")
    for p in (p_txt, p_csv, p_xlsx, p_iss, p_log):
        print("  ", p)


if __name__ == "__main__":
    main()
