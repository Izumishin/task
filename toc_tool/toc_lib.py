# -*- coding: utf-8 -*-
"""総目次Excel自動化のコアライブラリ。

構成:
  build_page_map(pdf_paths)      誌面ページ番号 -> (PDFパス, 0始まりページ索引)
  build_reading_dict(excel)      漢字姓名 -> よみ（既存Excel全行から。前年分含む）
  section_b_strings(excel)       区分キー -> B列の正確な文字列（●＋全角詰め）
  extract_article(page_text)     本文先頭ページ -> 記事情報（著者漢字/ローマ字/表題ほか）
  parse_toc(page_text)           目次ページ -> 記事の並び（区分/表題行/頁/筆頭著者）
  month_block_bounds(ws, month)  対象月データ行の開始・終了行
  write_month(...)               対象月ブロックを新データで置換（書式・数式・括弧を再現）

「毎月Claudeに依頼する」運用を前提に、機械的に確実な処理はここで完結させ、
判断が要る箇所（読みの長音、特集テーマ行の扱い等）は uncertain フラグと
レポートで表面化させて人／Claudeの確認に回す。
"""
import re
import copy
import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import PatternFill

from romaji import romaji_fullname_to_reading

IDEO_SP = "　"  # 全角スペース
YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# TOC/本文に現れる区分表記 -> 区分キー（正規化）
SECTION_ALIASES = {
    "巻頭言": "巻頭言", "特集": "特集", "原著論文": "原著論文",
    "原著": "原著論文", "基礎講座": "基礎講座", "連載": "連載",
    "文献抄録": "文献抄録", "書評": "書評", "学会NEWS": "学会NEWS",
    "学会News": "学会NEWS", "調査報告": "調査報告", "症例報告": "症例報告",
    "短報": "短報", "総説": "総説", "Series": "Series",
}


# ---------------------------------------------------------------------------
# PDF: ページマップ
# ---------------------------------------------------------------------------
def build_page_map(pdf_paths):
    """各PDFの各ページ先頭に印字された誌面ページ番号を読み、
    {誌面ページ番号(int): (pdf_path, page_index)} を返す。

    PDFが記事ごとにバラバラでも、何冊に分かれていても対応可能。
    同じ番号が複数ある場合は最初に見つかったものを採用。
    """
    page_map = {}
    docs = {}
    for path in pdf_paths:
        doc = fitz.open(path)
        docs[path] = doc
        for i in range(doc.page_count):
            first = doc[i].get_text().split("\n", 1)[0].strip()
            m = re.match(r"^(\d{1,4})$", first)
            if m:
                num = int(m.group(1))
                page_map.setdefault(num, (path, i))
    return page_map, docs


def page_text(page_map, docs, num):
    """誌面ページ番号のテキストを返す（無ければ None）。"""
    if num not in page_map:
        return None
    path, idx = page_map[num]
    return docs[path][idx].get_text()


# ---------------------------------------------------------------------------
# Excel: 既存データから読み辞書と区分文字列を作る
# ---------------------------------------------------------------------------
def build_reading_dict(ws, name_col=4, reading_col=9, max_row=None):
    """D列(著者漢字) -> I列(よみ) の辞書を既存全行から作る。

    同じ漢字名で複数の読みが出た場合は最頻を採用（表記ゆれ対策）。
    """
    from collections import Counter, defaultdict
    counts = defaultdict(Counter)
    max_row = max_row or ws.max_row
    for r in range(2, max_row + 1):
        name = ws.cell(row=r, column=name_col).value
        reading = ws.cell(row=r, column=reading_col).value
        if name and reading and isinstance(name, str) and isinstance(reading, str):
            counts[_norm_name(name)][reading.strip()] += 1
    return {k: v.most_common(1)[0][0] for k, v in counts.items()}


def _norm_name(s):
    """著者名の照合キー：空白（半角/全角）を除去。"""
    return s.replace(" ", "").replace(IDEO_SP, "").strip()


def section_b_strings(ws, max_row=None):
    """区分キー -> B列の正確な文字列 の辞書を既存全行から作る。"""
    result = {}
    max_row = max_row or ws.max_row
    for r in range(2, max_row + 1):
        b = ws.cell(row=r, column=2).value
        if not b or not isinstance(b, str):
            continue
        # "●　特　　集" -> コア文字 "特集"
        core = b.lstrip("●").replace(IDEO_SP, "").replace(" ", "").strip()
        # 増刊号など複数行のものは1行目のみで判定
        core = core.split("\n")[0]
        key = SECTION_ALIASES.get(core, core)
        # 短い（＝通常の区分見出し）ものを優先採用
        if key not in result or len(b) < len(result[key]):
            result[key] = b
    return result


def make_b_string(section_key, b_map):
    """区分キーからB列文字列を得る。未知なら簡易生成（要確認）。"""
    if section_key in b_map:
        return b_map[section_key], False
    return f"●{IDEO_SP}{section_key}", True


# ---------------------------------------------------------------------------
# 本文先頭ページの解析
# ---------------------------------------------------------------------------
_END_ANCHORS = re.compile(r"(抄\s*録|老年精神医学雑誌\s*\d+\s*[：:])")


def extract_article(text):
    """本文先頭ページのテキストから著者情報を抽出する。

    返り値 dict:
      kanji_authors : [漢字著者名, ...]（誌面の表示順）
      romaji_authors: [ローマ字著者名, ...]（脚注出現順、順序は表示順と一致しない場合あり）
      page_range    : (start, end) or None
    抽出できない項目は None/空。
    """
    lines = [ln.rstrip() for ln in text.split("\n")]

    page_range = None
    m = re.search(r"老年精神医学雑誌\s*\d+\s*[：:]\s*(\d+)\s*[-–]\s*(\d+)", text)
    if m:
        page_range = (int(m.group(1)), int(m.group(2)))

    # ローマ字著者：脚注行（"Name, Name：所属" / "＊1 Name：所属"）から
    # '：' より前のローマ字人名をすべて拾う。
    romaji_authors = []
    for ln in lines:
        s = ln.strip()
        if "：" not in s and ":" not in s:
            continue
        head = re.split(r"[：:]", s, 1)[0]
        head = re.sub(r"^＊?\d*\s*", "", head).strip()
        # 脚注以外の『：』行（E-mail, Key words, URL 等）を除外
        low = head.lower()
        if any(bad in low for bad in ("e-mail", "email", "key word", "http", "@", "tel", "fax", "〒")):
            continue
        # ローマ字人名（アルファベットとカンマ・空白・記号）のみを対象
        if head and re.fullmatch(r"[A-Za-zÀ-ÿ,\.\-’'\s　]+", head):
            for nm in head.split(","):
                nm = nm.strip()
                # 人名は2語以上（Given Surname）。1語や略語はノイズとして除外
                if nm and nm.lower() not in ("et al", "et al.") and len(nm.split()) >= 2:
                    romaji_authors.append(nm)

    # 漢字著者：末尾アンカー（抄録/雑誌行）直前の、中点/・区切りの漢字行。
    kanji_authors = _find_kanji_author_line(lines)

    return {
        "kanji_authors": kanji_authors,
        "romaji_authors": romaji_authors,
        "page_range": page_range,
    }


_KANJI = r"一-鿿぀-ゟ゠-ヿ々〆ヶ"


def _find_kanji_author_line(lines):
    """著者漢字名の並びを推定して返す。

    誌面では「表題／副題／著者漢字（＊n付き）／抄録」の順。
    ＊n や余分なトークンを除き、中点『・』区切りの人名列を復元する。
    """
    # アンカー行の位置
    anchor = None
    for i, ln in enumerate(lines):
        if _END_ANCHORS.search(ln):
            anchor = i
            break
    if anchor is None:
        anchor = len(lines)

    # アンカー直前をさかのぼり、著者行群（漢字＋・＋＊n）を集める。
    # 著者行は「ほぼ漢字＋中点＋＊＋数字＋読点」で構成される。
    collected = []
    j = anchor - 1
    author_line_re = re.compile(rf"^[\s{_KANJI}・,，＊\*0-9　]+$")
    # まず著者行の塊を見つける
    while j >= 0:
        s = lines[j].strip()
        if not s:
            j -= 1
            continue
        if author_line_re.match(s) and re.search(rf"[{_KANJI}]", s):
            collected.insert(0, s)
            j -= 1
            # 直上も著者の続き（＊付き複数行）なら続行、そうでなければ止める
            continue
        break

    if not collected:
        return []

    joined = "".join(collected)
    # ＊n マーカー・空白・末尾読点を除去し、中点で分割
    joined = re.sub(r"＊?\*?\d+", "・", joined)  # ＊1 → 区切り扱い
    joined = joined.replace(IDEO_SP, "").replace(" ", "")
    joined = joined.replace("，", "・").replace(",", "・").replace("、", "・")
    names = [n for n in joined.split("・") if re.search(rf"[{_KANJI}]", n)]
    return names


# ---------------------------------------------------------------------------
# 目次ページの解析
# ---------------------------------------------------------------------------
_SECTION_WORDS = ["巻頭言", "特集", "原著論文", "原著", "調査報告", "症例報告",
                  "基礎講座", "連載", "文献抄録", "書評", "学会NEWS",
                  "短報", "総説", "Series"]
_ZK = str.maketrans("０１２３４５６７８９", "0123456789")


def _clean_sec(word):
    w = word.replace(IDEO_SP, "").replace(" ", "")
    return SECTION_ALIASES.get(w, w)


def _looks_author(line):
    """著者行らしさ：漢字＋中点＋『ほか』のみで、句読点や助詞を含まない短い行。"""
    s = line.strip().replace(IDEO_SP, "").replace(" ", "")
    s = re.sub(r"\d+$", "", s)  # 末尾頁
    s = s.replace("ほか", "").replace("・", "")
    if not s:
        return False
    return bool(re.fullmatch(rf"[{_KANJI}]+", s)) and len(s) <= 12 * 6


def parse_toc(text):
    """目次ページのテキストから記事の並びを返す。

    返り値: [dict]  各 dict:
        section      区分キー
        title        表題（副題は改行結合、目次スタイル）
        page         開始頁 int（取れなければ None）
        toc_author   目次の筆頭著者表記（"繁田雅弘" / "伏屋研二ほか" 等、無ければ ""）
        need_body    True なら本文から表題取得が必要（文献抄録/書評 等）
    見出しテーマ行（特集の総合タイトル等）は section とともに theme=True で返す。
    """
    raw = [ln for ln in text.split("\n")]
    lines = [ln.strip() for ln in raw]
    entries = []
    cur_section = None
    i = 0
    n = len(lines)

    # 末尾の非記事（学会入会案内/投稿規定/目次/英字）以降は無視
    def is_noise(s):
        return (not s or s in ("目 次", "目次") or
                s.startswith("Japanese Journal") or
                s.startswith("Vol.") or
                "学会入会案内" in s or "投稿規定" in s or "編集後記" in s)

    pending_title = []   # 表題行の蓄積
    pending_page = None
    pending_section = None
    theme_pending = None  # 特集テーマ（著者なし見出し）

    def flush(author):
        nonlocal pending_title, pending_page, pending_section
        if pending_section and pending_title:
            title = "\n".join(t for t in pending_title if t.strip())
            entries.append({
                "section": pending_section,
                "title": title,
                "page": pending_page,
                "toc_author": author,
                "need_body": pending_section in ("文献抄録", "書評") and not pending_title,
            })
        pending_title = []
        pending_page = None

    while i < n:
        s = lines[i]
        if is_noise(s):
            i += 1
            continue

        # 行頭に頁番号＋区分（領域B）: "73  原著論文  タイトル..."
        m = re.match(rf"^(\d+)[\s　]+({'|'.join(_SECTION_WORDS)})[\s　]+(.*)$", s.translate(_ZK))
        if m:
            flush("")  # 前の記事を確定（著者は後続行で拾えなかった場合空）
            pending_section = _clean_sec(m.group(2))
            pending_page = int(m.group(1))
            rest = m.group(3).strip()
            pending_title = [rest] if rest else []
            i += 1
            continue

        # 行頭 "頁　区分"（タイトルが次行のこともある）
        m2 = re.match(rf"^(\d+)[\s　]+({'|'.join(_SECTION_WORDS)})\s*$", s.translate(_ZK))
        if m2:
            flush("")
            pending_section = _clean_sec(m2.group(2))
            pending_page = int(m2.group(1))
            pending_title = []
            i += 1
            continue

        # 領域A 巻頭言行: "3　巻頭言　自由と医療安全"
        m3 = re.match(rf"^(\d+)[\s　]+(巻頭言)[\s　]+(.*)$", s.translate(_ZK))
        if m3:
            flush("")
            pending_section = "巻頭言"
            pending_page = int(m3.group(1))
            pending_title = [m3.group(3).strip()]
            i += 1
            continue

        # 単独の区分見出し（"特集" など、領域A）
        if _clean_sec(s) in [ _clean_sec(w) for w in _SECTION_WORDS ] and len(s) <= 8:
            flush("")
            cur_section = _clean_sec(s)
            pending_section = cur_section
            i += 1
            continue

        # 著者行 + 末尾頁（領域A: "繁田雅弘　 5" / "伏屋研二ほか　13"）
        m4 = re.match(rf"^([{_KANJI}・]+ほか|[{_KANJI}・]+)[\s　]+(\d+)\s*$", s.translate(_ZK).replace(IDEO_SP, " ").strip())
        if m4 and pending_section:
            flush(m4.group(1))
            if pending_section != "巻頭言":
                pending_section = cur_section or pending_section
            continue_section = True
            i += 1
            # 領域Aでは特集内で pending_section を維持
            pending_section = cur_section if cur_section else pending_section
            continue

        # 著者のみの行（領域B: 頁は区分行で既知）
        if _looks_author(s) and pending_section and pending_title:
            flush(s)
            i += 1
            continue

        # それ以外は表題（副題含む）の一部として蓄積
        if pending_section is not None:
            # 副題の "――" や本文改行を保持
            if s:
                pending_title.append(s)
        i += 1

    flush("")
    # 特集の総合テーマ処理などは呼び出し側 / Claude 確認に委ねる
    return entries


# ---------------------------------------------------------------------------
# 読みの解決
# ---------------------------------------------------------------------------
def resolve_reading(kanji_name, romaji_name, reading_dict):
    """著者の読みを決める。返り値 (reading, uncertain)。

    優先: ①既存Excel辞書（漢字一致） → ②ローマ字変換（要確認）。
    """
    key = _norm_name(kanji_name)
    if key in reading_dict:
        return reading_dict[key], False
    if romaji_name:
        r, u = romaji_fullname_to_reading(romaji_name)
        return r, True  # 辞書に無い新規著者は必ず要確認
    return "", True


# ---------------------------------------------------------------------------
# Excel: 月ブロックの境界
# ---------------------------------------------------------------------------
def _separator_label(month):
    return f"{month}月号目次"


def month_block_bounds(ws, month):
    """対象月のデータ行 [start, end]（両端含む）を返す。

    月区切り行は G列に "N月号目次"（全角数字表記ゆれあり）。
    1月は区切り行が無く、ヘッダ行(1)の直後から始まる。
    """
    sep_rows = {}  # 正規化月ラベル -> 区切り行
    zk = str.maketrans("０１２３４５６７８９", "0123456789")
    for r in range(1, ws.max_row + 1):
        g = ws.cell(row=r, column=7).value
        if isinstance(g, str) and "月号目次" in g:
            label = g.translate(zk)
            m = re.match(r"(\d+)月号目次", label)
            if m:
                sep_rows[int(m.group(1))] = r
        if isinstance(g, str) and "増刊号" in g and "目次" in g:
            sep_rows.setdefault("zoukan", r)

    if month == 1:
        start = 2
    else:
        if month not in sep_rows:
            raise ValueError(f"{month}月の区切り行が見つかりません")
        start = sep_rows[month] + 1

    # 終了：次の区切り行の1つ前
    later = [row for key, row in sep_rows.items()
             if isinstance(key, int) and key > month]
    later += [row for key, row in sep_rows.items() if key == "zoukan"]
    if month != 1:
        later = [row for row in later if row > sep_rows[month]]
    end = (min(later) - 1) if later else ws.max_row
    return start, end


# ---------------------------------------------------------------------------
# Excel: 書き込み
# ---------------------------------------------------------------------------
def _copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.alignment = copy.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy.copy(src_cell.protection)


def _renumber_split_formulas(ws):
    """全行の J/K（姓・名分割数式）を現在の行番号で書き直す。

    行の挿入/削除で数式内の行参照がずれるため、最後に一括修復する。
    I列に値がある行だけ対象。
    """
    for r in range(2, ws.max_row + 1):
        i_val = ws.cell(row=r, column=9).value
        j = ws.cell(row=r, column=10)
        k = ws.cell(row=r, column=11)
        has_formula = (isinstance(j.value, str) and j.value.startswith("=")) or \
                      (isinstance(k.value, str) and k.value.startswith("="))
        if i_val or has_formula:
            j.value = f'=LEFT(I{r},FIND("{IDEO_SP}",I{r},1)-1)'
            k.value = (f'=MID(I{r},FIND("{IDEO_SP}",I{r},1)+1,'
                       f'LEN(I{r})-FIND("{IDEO_SP}",I{r},1)+1)')


def write_month(ws, month, rows, b_map):
    """対象月のデータ行を rows で置換する。

    rows: [dict] 各dict は
        section   区分キー（例 "特集"）
        title     C列表題（改行は "\n"）
        author    D列著者漢字（空可）
        page      E列頁（int/空）
        reading   I列よみ（空可）
        uncertain 真ならI列を黄色に
        note      任意メモ（レポート用）
    返り値: 追加行数と、黄色にしたセルの (row, 内容) 一覧。
    """
    start, end = month_block_bounds(ws, month)
    old_count = end - start + 1
    new_count = len(rows)

    # テンプレート行（書式コピー元）：置換前ブロックの先頭データ行。
    # 空ブロックに備え、無ければ2行目を使う。
    tmpl_row = start if ws.cell(row=start, column=2).value else 2
    tmpl_style = {c: ws.cell(row=tmpl_row, column=c) for c in range(1, 12)}
    tmpl_height = ws.row_dimensions[tmpl_row].height

    # 行数調整（挿入/削除）
    if new_count > old_count:
        ws.insert_rows(start, new_count - old_count)
    elif new_count < old_count:
        ws.delete_rows(start, old_count - new_count)

    yellow_cells = []
    for idx, row in enumerate(rows):
        r = start + idx
        # 書式をテンプレートからコピー
        for c in range(1, 12):
            _copy_style(tmpl_style[c], ws.cell(row=r, column=c))
        ws.row_dimensions[r].height = tmpl_height

        # 値
        ws.cell(row=r, column=1).value = None  # A列（共著順）は転記しない
        b_str, b_unknown = make_b_string(row["section"], b_map)
        ws.cell(row=r, column=2).value = b_str
        ws.cell(row=r, column=3).value = row.get("title") or None
        ws.cell(row=r, column=4).value = row.get("author") or None
        ws.cell(row=r, column=5).value = row.get("page") if row.get("page") else None
        ws.cell(row=r, column=6).value = "（"
        ws.cell(row=r, column=7).value = month
        ws.cell(row=r, column=8).value = "）"
        ws.cell(row=r, column=9).value = row.get("reading") or None
        # J/K は後段でまとめて付与

        if row.get("uncertain") or b_unknown:
            ws.cell(row=r, column=9).fill = YELLOW
            if b_unknown:
                ws.cell(row=r, column=2).fill = YELLOW
            yellow_cells.append((r, row.get("author", ""), row.get("reading", "")))

    _renumber_split_formulas(ws)
    return new_count, yellow_cells
